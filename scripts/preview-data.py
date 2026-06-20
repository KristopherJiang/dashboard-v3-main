#!/usr/bin/env python3
"""
读取 materials/ 目录下所有 xlsx 文件的表头 + 前 10 行数据
输出到终端，方便 AI 快速理解数据结构
"""

import os
import sys

try:
    import openpyxl
except ImportError:
    os.system("pip3 install openpyxl -q")
    import openpyxl

MATERIALS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "materials")

def read_xlsx_preview(filepath, max_rows=10):
    """读取 xlsx 文件的表头 + 前 N 行"""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*60}")
        print(f"文件: {os.path.basename(filepath)}")
        print(f"Sheet: {sheet_name}")
        print(f"总行数: {ws.max_row}")
        print(f"总列数: {ws.max_column}")
        print(f"{'='*60}")

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows + 1:  # +1 for header
                break
            if i == 0:
                print(f"\n【表头】{list(row)}")
                print(f"{'-'*60}")
            else:
                print(f"【第{i}行】{list(row)}")

    wb.close()

def main():
    if not os.path.exists(MATERIALS_DIR):
        print(f"错误: {MATERIALS_DIR} 目录不存在")
        return

    xlsx_files = [f for f in os.listdir(MATERIALS_DIR) if f.endswith('.xlsx')]

    if not xlsx_files:
        print("没有找到 xlsx 文件")
        return

    print(f"找到 {len(xlsx_files)} 个 xlsx 文件\n")

    for f in sorted(xlsx_files):
        filepath = os.path.join(MATERIALS_DIR, f)
        read_xlsx_preview(filepath, max_rows=10)
        print("\n")

if __name__ == "__main__":
    main()
